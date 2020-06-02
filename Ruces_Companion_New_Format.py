'''
# NOTES:
#
# KIVY LIBRARY INSTALL AND USE
# https://kivy.org/doc/stable/installation/installation-windows.html#install-win-dist
#
# ISSUE WITH 'APP': https://stackoverflow.com/questions/28424188/no-module-named-kivy-app
#
# TKINTER APPLICATION WINDOWS
# https://effbot.org/tkinterbook/tkinter-application-windows.htm
#
# ACTIVATE THE KIVY ENVIRONMENT, REQUIRED!!!
# kivy_venv\Scripts\activate
#
# READING AND WRITING EXCEL FILES
# http://zetcode.com/articles/openpyxl/
#
# SELECTING A FILE
# https://www.youtube.com/watch?v=H71ts4XxWYU
#
# ACCESSING ONLINE FILES
# https://stackoverflow.com/questions/7243750/download-file-from-web-in-python-3
#
# CONVERTING INPUT 
# *To change worksheet: sheet_required = workbook_variable_name.sheetnames[n], where 'n' is an index starting with 0
#  Do get sheets by name: sheet_required = workbook_variable_name[sheetname]
#
# MOBILE
# https://stackoverflow.com/questions/101754/is-there-a-way-to-run-python-on-android
#
# *Use user entered variable name
#  https://stackoverflow.com/questions/29824194/setting-user-input-to-variable-name
#
# *Remember to keep the logic and GUI decoupled, so if needed it may be transferred
#
# *It may be a good idea to build the GUI and logic in separate ".py" files
#
# ADDING SCREENS AS NEEDED
#
# https://kivy.org/doc/stable/api-kivy.uix.screenmanager.html?highlight=screen
#
# CREATING A CUSTOM DICTIONARY
# https://www.geeksforgeeks.org/python-add-new-keys-to-a-dictionary/
#
# PROJECT DESCRIPTION:
#
# Create a program to complement, and possibly replace RUCES. The program will eventually be able to scan through a
# specified excel workbook, containing all of the usual table entries for and IPM or LR job, and compile a complete
# RUCES like cost estimation for cost and labour etc. The program should also allow for addition of extra parts such as
# Neutral Brackets and armor rods
#
# BUILD OBJECTIVES
#
# (SECONDARY) Add option to build the IPM sheet through the program itself [ON HOLD]
#
# Overhauld to be used on a server instead of local files [ON HOLD]
#
# Add functionality for labour and direct costs, asset allocation [ON HOLD]
#
# Overhaul database and change code to match [COMPLETE FOR EXCEL, ADAPT SQL]
#
# Handle Exceptions [IP]
#
# Update the CD database [ONGOING]
#
# Automate the process of determining pole type and update the cd_cat dictionary [COMPLETE]
#
# Find a way to determine which sheet contains the CD numbers and IPM info [ON HOLD]
#
# Scan a given IPM/LR job excel file and automatically determine the needed parts [COMPLETE FOR EXCEL, ADAPT SQL]
#
# Write the parts to an excel sheet [COMPLETE]
#
# Take in a quantity and multiple CDs and run similar logic [COMPLETE]
#
# Have the program scan through the database and return the part count for a given CD number [COMPLETE]
#
'''




#Import the Excel spreadsheet library
import openpyxl
import sys
import os
import tkinter as tk
from tkinter import filedialog

#Create the tkinter root
root = tk.Tk()
root.withdraw()

cd_path = "cd_database.xlsx"
cd_data = None

def  retrieveDatabase():
    #Specify the path to the cd database file
    global cd_path
    global cd_data
    

    #Load the cd database workbook into a variable and give the option to locate the database in case filename was changed
    try:
        cd_data = openpyxl.load_workbook(cd_path)
    except FileNotFoundError as fnf_error:
        search = input("The main database could not be located, this is NOT your IPM spreadsheet. Would you like to point to the database location? (Y - Yes, N - No): ")
        while True:
            if search.upper() == 'Y':
                cd_path = filedialog.askopenfilename()
                cd_data = openpyxl.load_workbook(cd_path)
                break
            elif search.upper() == 'N':
                input("Unable to continue without database, contact administrator. Press Enter to exit.")
                quit()
            else:
                search = input("Please enter 'Y' or 'N' for 'Yes' or 'No': ")
    else:
        print("Database found with name: ", cd_path)

    return cd_path
        
        
    
        

#Variable for the IPM job workbook
ipm_data = None

#global directory of parts and their quantitites, updated whenever the part has a non zero value in cd_database
#Match these with the header names in the cd_database workbook
part_list = {}

pole_list = {}

ipm_list = {}

#Dictionary of all the required CD insances
cd_list = {}
cd_counter = {}
cd_list_names = {}
missing_cd = {}

#Variable for IPM column header
ipm_col_name = None

#Ipm loaded variable
ipm_loaded = False

#Class for each CD number
class PoleIPM:
    
    def __init__(self, pole, CD, conductor, name):

        self.pole = pole
        self.CD = CD
        self.conductor = conductor
        self.name = name

    ##Function for scanning through cd numbers and getting part quantities
    def scanCD(self):

        if self.CD == "":
            missing_cd.update({self.name: self})
            return
        
        global cd_data
        global part_list
        
        print("scanning...")
        ws_cd = cd_data[self.CD.split('-')[0]]
        ws_parts = cd_data["PARTS"]
        cd_row = None

        for col in ws_cd.iter_cols(min_row = 2, min_col = 1, max_row = ws_cd.max_row, max_col = 1):
            for cell in col:
                if str(cell.value) == str(self.CD):
                    print("MATCH")
                    cd_row = cell.row
                    break
                elif cell.value is None:
                    return
        if cd_row is None:
            return
        for row in ws_cd.iter_rows(min_row = cd_row, min_col = 2, max_row = cd_row, max_col = ws_cd.max_column):
            for cell in row:
                if cell.value is None:
                    break

                if cell.value == 0:
                    pass
                #Case for part with specific conductor
                elif type(part_list[str(ws_cd.cell(row = 1, column = cell.column).value)]) is dict and '*' not in str(ws_cd.cell(row = 1, column = cell.column).value):
                    for col in ws_parts.iter_cols(min_row = 2, min_col = 1, max_row = ws_parts.max_row, max_col = 1):
                        for i in col:
                            if str(i.value) in str(ws_cd.cell(row = 1, column = cell.column).value):
                                for cur_part in ws_parts.iter_rows(min_row = i.row, min_col = 2, max_row = i.row, max_col = ws_parts.max_column):
                                    for entry in cur_part:
                                        if entry.value is None:
                                            print("Part for this conductor could not be found...")
                                            break
                                        elif str(entry.value) in str(self.conductor):
                                            if entry.value in part_list[i.value]:
                                                part_list[str(i.value)][entry.value] += int(cell.value)
                                            else:
                                                part_list[i.value].update({str(entry.value): 1})
                                            break
                #Case for generic entry
                else:
                    part_list[str(ws_cd.cell(row = 1, column = cell.column).value)] += int(cell.value)      

#Function for retrieving the IPM spreadsheet
def openIPM():
    
    global ipm_data
    global ipm_loaded
    
    print('opening')
    ipm_path = filedialog.askopenfilename()
    print('done')
    
    
    try:
        ipm_data = openpyxl.load_workbook(ipm_path)
        ipm_loaded = True
    except:
        return 'IPM Not Loaded'
        ipm_loaded = False

    return ipm_path

#Function for receiving a lsit of IPM numbers
def getIPM(ipm_column_entered):

    global ipm_data
    global ipm_col_name
    global ipm_list
    ws = ipm_data.active
    ipm_col_name = ipm_column_entered

    for row in ws.iter_rows(min_row = 1, min_col = 1, max_row = 1, max_col = ws.max_column):
        for cell in row:
            if str(cell.value.upper()) == ipm_col_name.upper():
                ipm_col = cell.column
                break
                
    for col in ws.iter_cols(min_row = 2, min_col = ipm_col, max_row = ws.max_row, max_col = ipm_col):
        for cell in col:
            if cell.value is not None:
                ipm_list.update({str(cell.value): None})
    

#Build the ipm object list
def buildIPM(pole_header, cd_header, conductor_header):

    global ipm_list
    for ipm in ipm_list:
        ipm_vars = makeIPM(str(ipm), str(pole_header), str(cd_header), str(conductor_header))
        
        globals()[str(ipm)] = PoleIPM(ipm_vars[0], ipm_vars[1], ipm_vars[2], str(ipm))
        ipm_list[ipm] = globals()[str(ipm)]

#Function for creating IPM objects
def  makeIPM(ipm, pole_header, cd_header, conductor_header):

    global ipm_data
    global ipm_col_name
    ws = ipm_data.active

    for row in ws.iter_rows(min_row = 1, min_col = 1, max_row = 1, max_col = ws.max_column):
        for cell in row:
            if cell.value == pole_header:
                pole_col = cell.column
            elif cell.value == cd_header:
                cd_col = cell.column
            elif cell.value == conductor_header:
                conductor_col = cell.column

    for row in ws.iter_rows(min_row = 1, min_col = 1, max_row = 1, max_col = ws.max_column):
        for cell in row:
            if cell.value.upper() == ipm_col_name.upper():
                ipm_col = cell.column
                break
    for col in ws.iter_cols(min_row = 2, min_col = ipm_col, max_row = ws.max_row, max_col = ipm_col):
        for cell in col:
            if str(cell.value) == str(ipm):
                ipm_row = cell.row
                break
    for row in ws.iter_rows(min_row = ipm_row, min_col = 1, max_row = ipm_row, max_col = ws.max_column):
        for cell in row:
            if cell.value is None:
                if cell.column == pole_col:
                    pole = None
                elif cell.column == cd_col:
                    cd = ""
                elif cell.column == conductor_col:
                    conductor = ""
            elif cell.column == pole_col:
                pole = str(cell.value)
            elif cell.column == cd_col:
                if "CD" in cell.value:
                    cd = str(cell.value)
                else:
                    cd = "CD" + str(cell.value)
            elif cell.column == conductor_col:
                conductor = ""
                for char in cell.value:
                    try:
                        x = int(char)
                        conductor = cell.value[cell.value.index(char):] #find the first number and start the conductor name there
                        if "3/13" in conductor and "ST" not in conductor.upper():
                            conductor += "ST"
                        break
                    except:
                        pass
            elif cell.column == pole_col:
                pole = str(cell.value)
    try:
        return pole, cd, conductor
    except TypeError:
        return pole, cd, conductor
    
#Create a list of all parts
def makePartList():

    #Retrieve the global parts list dictionary
    global part_list
    global cd_data

    #Activate the parts count sheet which contains all parts as headers
    sheet = cd_data['PARTS']

    #Scan the first column and store the part names in the part list tuple
    for col in sheet.iter_cols(min_row = 2, min_col = 1, max_row = sheet.max_row, max_col = 1):
        for cell in col:
            if cell.value is None:
                break
            if sheet.cell(cell.row, cell.column+1).value is not None:
                part_list.update({cell.value: {}})
            else:
                part_list.update({cell.value: 0})
    print(part_list)


#Create a list of all the poles
def makePoleList():
    global pole_lsit
    global cd_data

    sheet = ipm_data.active

    for row in sheet.iter_rows(min_row = 1, min_col = 1, max_row = 1, max_col = sheet.max_column):
        for cell in row:
            if "POLE" in cell.value.upper():
                pole_col = cell.column
                break
    for col in sheet.iter_cols(min_row = 2, min_col = pole_col, max_row = sheet.max_row, max_col = pole_col):
        for cell in col:
            if cell.value in pole_list:
                pole_list[cell.value] += 1
            else:
                pole_list.update({cell.value: 1})


def scan():
    
    for ipm in ipm_list:
        print(ipm)
        ipm_list[ipm].scanCD()

#function for writing the parts to their respective sheet
##def writeToSheet():
##
##    #Access the parts count worksheet
##    ws = cd_data['PARTS COUNT']
##    
##
##    #Scan the workbook, write the part counts to the proper columns
##    try:
##        for col in ws.iter_cols(min_row = 2, min_col = 1, max_row = ws.max_row, max_col =1 ):
##            for cell in col:
##                ws.cell(row=cell.row, column=cell.column+1).value = part_list[str(cell.value)]
##    except KeyError:
##        print("\nWARNING!: None value received...check spreadsheet.")
##
##    #Create a new excel file for the parts list
##    p_list_book = openpyxl.Workbook()
##    sheet = p_list_book.active
##
##    
##    for col in ws.iter_cols(min_row = 2, min_col = 1, max_row = ws.max_row, max_col =2 ):
##        for cell in col:
##            sheet.cell(row=cell.row, column=cell.column).value = cell.value
##
##    #Ask the user where they would like the file saved
##    #Save the book with the appropriate file extension
##    while True:
##        p_list_path = filedialog.asksaveasfilename()
##        try:
##            if  'xlsx' in p_list_path: 
##                p_list_book.save(p_list_path)
##            else:
##                p_list_book.save(p_list_path + '.xlsx')
##            break
##                
##        except PermissionError:
##            input("Please close the parts list file and try again. Press Enter when ready...")




#Main Code
def main():

    global cd_data

    #Access the cd database
    retrieveDatabase()

    #Open the workbook containing all the IPM data
    openIPM()

    #Create a dictionary of all the parts
    makePartList()

    #Create a dictionary of all the poles
    makePoleList()
    
    #Retrieve the list of CD numbers from the IPM job
    getIPM()
    
    #Create instances of all the IPM numbers
    pole_header = input("Enter the name of your pole column: ").upper()
    cd_header = input("Enter the name of your CD# column: ").upper()
    conductor_header = input("Enter the name of yout conductor column: ").upper()
    for ipm in ipm_list:
        ipm_vars = makeIPM(str(ipm), str(pole_header), str(cd_header), str(conductor_header))
        
        globals()[str(ipm)] = PoleIPM(ipm_vars[0], ipm_vars[1], ipm_vars[2], str(ipm))
        ipm_list[ipm] = globals()[str(ipm)]


    #Scan for the appropriate parts for each IPM

    for ipm in ipm_list:
        print(ipm)
        ipm_list[ipm].scanCD()
    
    print(part_list)     
##    writeToSheet()

    #Save the database
    while True:
        try:
            cd_data.save(cd_path)
            break
        except PermissionError:
            input("Please close the database and try again. Press Enter when ready.\n")

#Run the program
if __name__ == "__main__":
    main()

    
